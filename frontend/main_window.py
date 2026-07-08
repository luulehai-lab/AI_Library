# Tên file: frontend/main_window.py
# CHANGELOG:
# - 2024-08-06: [NÂNG CẤP] Thêm slot on_thumbnail_refresh_requested và xử lý logic xoay ảnh trong on_processing_finished. (AI Studio)
# - 2024-08-06: [NÂNG CẤP] Tích hợp DashboardWidget vào giao diện chính và thêm nút chuyển đổi. (AI Studio)
# - 2024-08-06: [VÁ LỖI] Thêm slot on_header_filter_applied và kết nối với tín hiệu từ BookTableWidget để xử lý logic lọc, sửa lỗi crash. (AI Studio)
# - 2024-08-06: [TỐI ƯU] Cập nhật luồng xử lý AI và thủ công để tạo thumbnail sau khi thêm sách, giúp tăng tốc độ tải. (AI Studio)
# - 2024-08-06: [VÁ LỖI] Chuẩn hóa việc truyền dữ liệu cho hàm add_book, đảm bảo luồng AI và thủ công đều dùng key là tên cột DB. (AI Studio)
# - 2024-08-06: [NÂNG CẤP] Thêm nút "Thêm sách thủ công" và logic để mở dialog, xử lý dữ liệu trả về, và lưu sách mới. (AI Studio)
# - 2024-08-06: [TÁI CẤU TRÚC] Tách toàn bộ logic và UI của bảng hiển thị sách ra BookTableWidget. MainWindow giờ chỉ quản lý và kết nối tín hiệu. (AI Studio)
# - 2024-08-06: [TÁI CẤU TRÚC] Tách toàn bộ logic và UI của khu vực chat ra ChatWidget riêng biệt. MainWindow chỉ còn nhiệm vụ khởi tạo và thêm widget này. (AI Studio)
# - 2024-08-06: [NÂNG CẤP] Cập nhật phiên bản lên 1.7.0. (AI Studio)
# - 2024-08-05: [NÂNG CẤP] ExcelExportWorker: Thêm tiêu đề lớn, xóa text đường dẫn ảnh, wrap text, định dạng Table, chỉnh độ rộng cột Review. (AI Studio)
# - 2024-08-05: [VÁ LỖI] Thêm import QObject, pyqtSignal, QUrl còn thiếu. Sửa logic mở file bằng QUrl.fromLocalFile. (AI Studio)
# - 2024-08-05: [NÂNG CẤP] Thêm tính năng xuất Excel (chọn/toàn bộ) và sắp xếp cột bằng cách nháy chuột trái vào header. (AI Studio)
# - 2024-08-04: [TỐI ƯU] Triển khai debouncing cho ô tìm kiếm để tránh lag. Cập nhật status bar để hiển thị thống kê khi lọc. (AI Studio)
# - 2024-08-04: [NÂNG CẤP] Tái cấu trúc Gallery View để phân biệt nháy đúp vào ảnh (phóng to) và chữ (xem chi tiết). Khôi phục dialog chi tiết. (AI Studio)
# - 2024-08-04: [TÁI CẤU TRÚC] Thay đổi logic nút "Thêm sách" để tự động quét và xử lý thư mục 'images_input'. (AI Studio)
# - 2024-08-03: [VÁ LỖI] Sửa lỗi QThread bị hủy sớm bằng cách gán parent (self) cho QThread khi khởi tạo. (AI Studio)
# - 2024-08-03: [NÂNG CẤP] Triển khai xử lý hàng loạt ảnh thông qua hệ thống hàng đợi (queue). (AI Studio)
# - 2024-08-03: [TÁI CẤU TRÚC] Chuyển sang Light Theme, bỏ dialog chi tiết, nháy đúp ảnh để phóng to, render Markdown, vá lỗi status bar. (AI Studio)
# - 2024-08-02: [NÂNG CẤP] Thêm hàng bộ lọc cho từng cột, triển khai menu ngữ cảnh và logic cho cột Status, cập nhật thanh trạng thái. (AI Studio)
# - 2024-08-01: [NÂNG CẤP] Cho phép chỉnh sửa trực tiếp trên bảng, thêm tổng giá tiền, tiêu đề thư viện, trạng thái đọc nâng cao và vá lỗi. (AI Studio)
# - 2024-07-31: [NÂNG CẤP] Triển khai banner đa ảnh, có thể cuộn, thêm/xóa và lưu trạng thái. (AI Studio)
# - 2024-07-31: [NÂNG CẤP] Tích hợp QSettings để lưu trạng thái UI, cho phép chọn banner, và phóng to ảnh bìa khi nháy đúp. (AI Studio)
# - 2024-07-31: [NÂNG CẤP] Thêm banner, tự động giãn dòng cho review, thanh cuộn cho dialog chi tiết, và hiển thị tổng số trang. (AI Studio)
# - 2024-07-30: [TÁI CẤU TRÚC] Tích hợp SQLite, giao diện dark mode, thanh tìm kiếm, chế độ xem lưới, và nhiều cải tiến UI/UX khác. (AI Studio)
# - 2024-07-29: [KHỞI TẠO] Xây dựng giao diện chính của ứng dụng. (AI Studio)

import os
import shutil
import markdown
import re
from datetime import datetime
from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QFileDialog, QProgressBar, QLabel, QMessageBox, QLineEdit, 
                             QStatusBar, QStackedWidget, QListWidget, QListWidgetItem,
                             QDialog, QFormLayout, QTextEdit, QScrollArea, QMenu, QSplitter)
from PyQt6.QtCore import QThread, Qt, QSettings, QTimer, QObject, pyqtSignal, QUrl
from PyQt6.QtGui import QPixmap, QAction, QDesktopServices
from backend.ai_service import AIServiceWorker
from backend.database import (get_filtered_books, add_book, get_library_stats, 
                              get_book_by_id)
from backend.rag_service import embed_and_store_book
from backend.utils import find_cover_image, create_thumbnail, get_thumbnail_path, rotate_image_file
from backend.data_model import get_ai_to_db_key_map
from config import APP_NAME, APP_VERSION, SETTINGS_ORGANIZATION, SETTINGS_APP_NAME
from .styles import STYLESHEET
from .widgets import ClickableLabel, GalleryItemWidget
from .components.chat_widget import ChatWidget
from .components.book_table_widget import BookTableWidget
from .components.add_book_dialog import AddBookDialog
from .components.dashboard_widget import DashboardWidget
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelExportWorker(QObject):
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, file_path, headers, db_columns, book_data):
        super().__init__()
        self.file_path = file_path
        self.headers = headers
        self.db_columns = db_columns
        self.book_data = book_data

    def run(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Thư viện sách"

            # 1. Tạo tiêu đề lớn
            ws.merge_cells('A1:Q1') # Giả sử bảng có khoảng 17 cột
            title_cell = ws['A1']
            title_cell.value = "THƯ VIỆN GIA ĐÌNH LÊ HẢI LƯU"
            title_cell.font = Font(size=20, bold=True, color="0000FF")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40

            # 2. Ghi tiêu đề cột (Hàng 2)
            ws.append(self.headers)
            
            # Định dạng tiêu đề cột
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 3. Ghi dữ liệu (Từ hàng 3)
            for book in self.book_data:
                row_data = []
                for col in self.db_columns:
                    if col == 'id': continue
                    # Nếu là cột cover_path, ta ghi chuỗi rỗng để chỉ hiện ảnh
                    if col == 'cover_path':
                        row_data.append("") 
                    else:
                        row_data.append(book.get(col, ""))
                ws.append(row_data)

            # 4. Chèn ảnh và Định dạng Cell
            cover_col_index = self.db_columns.index('cover_path')
            # Tìm vị trí cột Bìa sách trong Excel (1-based)
            try:
                excel_cover_col_idx = self.headers.index("Bìa sách") + 1
                cover_col_letter = get_column_letter(excel_cover_col_idx)
                ws.column_dimensions[cover_col_letter].width = 15
            except ValueError:
                cover_col_letter = 'A'

            # Tìm vị trí cột Review để chỉnh độ rộng
            try:
                review_col_idx = self.headers.index("Review") + 1
                ws.column_dimensions[get_column_letter(review_col_idx)].width = 50
            except ValueError:
                pass

            # Định dạng Wrap Text cho toàn bộ vùng dữ liệu
            last_row = len(self.book_data) + 2
            last_col_letter = get_column_letter(len(self.headers))
            data_range = ws[f"A3:{last_col_letter}{last_row}"]
            
            for row in data_range:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')

            # Chèn ảnh
            for i, book in enumerate(self.book_data):
                row_num = i + 3 # Hàng dữ liệu bắt đầu từ 3
                ws.row_dimensions[row_num].height = 95
                cover_path = book.get('cover_path')
                if cover_path and os.path.exists(cover_path):
                    try:
                        img = OpenpyxlImage(cover_path)
                        img.height = 120
                        img.width = 90
                        cell_address = f"{cover_col_letter}{row_num}"
                        ws.add_image(img, cell_address)
                    except Exception as img_e:
                        print(f"Không thể chèn ảnh {cover_path}: {img_e}")

            # 5. Tạo Table (Excel Table)
            tab = Table(displayName="BookTable", ref=f"A2:{last_col_letter}{last_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            wb.save(self.file_path)
            self.finished.emit(self.file_path)
        except Exception as e:
            self.error.emit(f"Lỗi khi xuất file Excel: {str(e)}")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.thread = None
        self.worker = None
        self.export_thread = None
        self.export_worker = None
        self.image_queue = []
        self.total_images_in_queue = 0
        self.active_filters = {}
        self.sort_column = None
        self.sort_order = Qt.SortOrder.AscendingOrder
        
        self.filter_timer = QTimer(self)
        self.filter_timer.setSingleShot(True)
        self.filter_timer.timeout.connect(self.trigger_filter)

        self.grand_total_stats = {}

        self.initUI()
        self.load_settings()
        self.load_books_from_db()

    def initUI(self):
        self.setWindowTitle(f"{APP_NAME} - v{APP_VERSION}")
        self.setGeometry(100, 100, 1800, 1000)
        self.setStyleSheet(STYLESHEET)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        self.setup_banner_area()
        self.layout.addWidget(self.banner_scroll_area)

        self.setup_control_bar()

        main_splitter = QSplitter(Qt.Orientation.Vertical)
        self.layout.addWidget(main_splitter)

        # Setup view container (Table, Gallery, Dashboard)
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0,0,0,0)
        self.stacked_widget = QStackedWidget()
        top_layout.addWidget(self.stacked_widget)
        main_splitter.addWidget(top_widget)

        # 1. BookTableWidget
        self.book_table_widget = BookTableWidget(self)
        self.stacked_widget.addWidget(self.book_table_widget)
        
        # 2. Gallery View
        self.setup_gallery_view()

        # 3. DashboardWidget (MỚI)
        self.dashboard_widget = DashboardWidget(self)
        self.stacked_widget.addWidget(self.dashboard_widget)

        # ChatWidget
        self.chat_widget = ChatWidget(self)
        main_splitter.addWidget(self.chat_widget)
        main_splitter.setSizes([800, 250])

        self.setup_status_area()
        
        # Kết nối tín hiệu từ các component con
        self.connect_signals()

    def setup_control_bar(self):
        """Thiết lập thanh công cụ điều khiển chính."""
        control_layout = QHBoxLayout()
        self.add_banner_button = QPushButton("Thêm ảnh bìa...")
        control_layout.addWidget(self.add_banner_button)

        self.clear_banner_button = QPushButton("Xóa ảnh bìa")
        control_layout.addWidget(self.clear_banner_button)

        self.library_title = QLabel("THƯ VIỆN GIA ĐÌNH LÊ HẢI LƯU")
        self.library_title.setObjectName("libraryTitleLabel")
        control_layout.addWidget(self.library_title, 0, Qt.AlignmentFlag.AlignCenter)
        control_layout.addStretch(1)

        self.add_manual_button = QPushButton("Thêm sách thủ công...")
        control_layout.addWidget(self.add_manual_button)

        self.upload_button = QPushButton("Xử lý thư mục Input")
        control_layout.addWidget(self.upload_button)
        
        self.export_button = QPushButton("Xuất Excel...")
        self.setup_export_menu()
        control_layout.addWidget(self.export_button)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Tìm kiếm chung...")
        control_layout.addWidget(self.search_input)

        self.view_toggle_button = QPushButton("Xem dạng Lưới")
        control_layout.addWidget(self.view_toggle_button)
        
        # Nút chuyển sang Dashboard
        self.dashboard_button = QPushButton("Dashboard")
        self.dashboard_button.clicked.connect(self.toggle_dashboard)
        control_layout.addWidget(self.dashboard_button)

        self.layout.addLayout(control_layout)

    def setup_status_area(self):
        """Thiết lập khu vực hiển thị trạng thái và thanh tiến trình."""
        self.status_label = QLabel("Sẵn sàng.")
        self.layout.addWidget(self.status_label)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.layout.addWidget(self.progress_bar)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

    def connect_signals(self):
        """Kết nối tất cả các tín hiệu và slot ở một nơi."""
        # Control bar signals
        self.add_banner_button.clicked.connect(self.add_banner_images)
        self.clear_banner_button.clicked.connect(self.clear_banner_images)
        self.add_manual_button.clicked.connect(self.show_add_manual_dialog)
        self.upload_button.clicked.connect(self.start_processing)
        self.search_input.textChanged.connect(self.apply_filters_debounced)
        self.view_toggle_button.clicked.connect(self.toggle_view)

        # BookTableWidget signals
        self.book_table_widget.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.book_table_widget.data_changed.connect(self.update_status_bar)
        self.book_table_widget.filter_triggered.connect(self.trigger_filter)
        self.book_table_widget.view_image_requested.connect(self.show_full_size_image)
        self.book_table_widget.status_message_changed.connect(lambda msg: self.status_bar.showMessage(msg, 3000))
        self.book_table_widget.header_filter_applied.connect(self.on_header_filter_applied)
        self.book_table_widget.thumbnail_refresh_requested.connect(self.on_thumbnail_refresh_requested)

    def on_thumbnail_refresh_requested(self, book_id):
        """Slot để làm mới thumbnail khi người dùng yêu cầu."""
        book = get_book_by_id(book_id)
        if book and book['cover_path']:
            create_thumbnail(book['cover_path'], book_id)
            self.status_bar.showMessage(f"Đã làm mới ảnh bìa cho sách ID {book_id}", 3000)
            self.trigger_filter() # Tải lại để hiển thị ảnh mới
        else:
            self.status_bar.showMessage(f"Không tìm thấy ảnh bìa để làm mới cho sách ID {book_id}", 3000)

    def on_header_filter_applied(self, column, value):
        """Slot để xử lý khi người dùng chọn một bộ lọc từ header."""
        if value is None:
            self.active_filters.pop(column, None)
        else:
            self.active_filters[column] = value
        self.trigger_filter()

    def show_add_manual_dialog(self):
        """Mở dialog để thêm sách thủ công và xử lý kết quả."""
        dialog = AddBookDialog(self)
        result = dialog.exec()

        if result == QDialog.DialogCode.Accepted:
            book_data = dialog.get_book_data()
            
            book_title = book_data.get('ten_sach')
            cover_path = find_cover_image(book_title)
            book_data['cover_path'] = cover_path
            
            try:
                book_id = add_book(book_data)
                if cover_path:
                    create_thumbnail(cover_path, book_id)
                
                embed_and_store_book(book_id)
                self.trigger_filter()
                self.status_bar.showMessage(f"Đã thêm thành công sách '{book_title}'", 5000)
            except Exception as e:
                self.show_error(f"Lỗi khi lưu sách mới: {e}")

    def setup_export_menu(self):
        menu = QMenu(self)
        action_export_selected = QAction("Xuất các sách đã chọn", self)
        action_export_selected.triggered.connect(self.export_selected_to_excel)
        menu.addAction(action_export_selected)

        action_export_all = QAction("Xuất toàn bộ thư viện", self)
        action_export_all.triggered.connect(self.export_all_to_excel)
        menu.addAction(action_export_all)
        
        self.export_button.setMenu(menu)
    
    def setup_banner_area(self):
        self.banner_scroll_area = QScrollArea()
        self.banner_scroll_area.setObjectName("bannerScrollArea")
        self.banner_scroll_area.setWidgetResizable(True)
        self.banner_scroll_area.setFixedHeight(150)
        self.banner_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.banner_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        self.banner_widget = QWidget()
        self.banner_layout = QHBoxLayout(self.banner_widget)
        self.banner_layout.setContentsMargins(0, 0, 0, 0)
        self.banner_layout.setSpacing(5)
        
        self.banner_scroll_area.setWidget(self.banner_widget)

    def add_banner_images(self):
        file_paths, _ = QFileDialog.getOpenFileNames(self, "Chọn ảnh bìa", "", "Image Files (*.png *.jpg *.jpeg)")
        if not file_paths: return

        settings = QSettings(SETTINGS_ORGANIZATION, SETTINGS_APP_NAME)
        current_paths = settings.value("bannerPaths", [], type=list)
        banner_storage_path = os.path.join("assets", "custom_banners")

        for path in file_paths:
            try:
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
                _, extension = os.path.splitext(path)
                new_filename = f"banner_{timestamp}{extension}"
                destination_path = os.path.join(banner_storage_path, new_filename)
                shutil.copy(path, destination_path)
                current_paths.append(destination_path)
            except Exception as e:
                self.show_error(f"Không thể sao chép file ảnh: {e}")
        
        settings.setValue("bannerPaths", current_paths)
        self.load_banner_images()

    def clear_banner_images(self):
        reply = QMessageBox.question(self, "Xác nhận", "Bạn có chắc muốn xóa tất cả ảnh bìa?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No: return

        settings = QSettings(SETTINGS_ORGANIZATION, SETTINGS_APP_NAME)
        paths_to_delete = settings.value("bannerPaths", [], type=list)
        
        for path in paths_to_delete:
            if os.path.exists(path):
                try: os.remove(path)
                except OSError as e: self.show_error(f"Không thể xóa file: {e}")

        settings.setValue("bannerPaths", [])
        self.load_banner_images()

    def load_banner_images(self):
        while self.banner_layout.count():
            child = self.banner_layout.takeAt(0)
            if child.widget(): child.widget().deleteLater()

        settings = QSettings(SETTINGS_ORGANIZATION, SETTINGS_APP_NAME)
        banner_paths = settings.value("bannerPaths", [], type=list)

        if not banner_paths:
            placeholder = QLabel("Thư Viện Sách Cá Nhân\n(Nhấn 'Thêm ảnh bìa...' để bắt đầu)")
            placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
            placeholder.setObjectName("bannerPlaceholder")
            self.banner_layout.addWidget(placeholder)
            return

        for path in banner_paths:
            if os.path.exists(path):
                pixmap = QPixmap(path)
                if not pixmap.isNull():
                    label = QLabel()
                    scaled_pixmap = pixmap.scaledToHeight(self.banner_scroll_area.height() - 20, Qt.TransformationMode.SmoothTransformation)
                    label.setPixmap(scaled_pixmap)
                    self.banner_layout.addWidget(label)
        self.banner_layout.addStretch(1)

    def on_header_clicked(self, logical_index):
        db_col = self.book_table_widget.db_columns[logical_index]
        if db_col in ["id", "cover_path"]: return

        if self.sort_column == db_col:
            self.sort_order = Qt.SortOrder.DescendingOrder if self.sort_order == Qt.SortOrder.AscendingOrder else Qt.SortOrder.AscendingOrder
        else:
            self.sort_column = db_col
            self.sort_order = Qt.SortOrder.AscendingOrder
        
        self.book_table_widget.table.horizontalHeader().setSortIndicator(logical_index, self.sort_order)
        self.trigger_filter()

    def setup_gallery_view(self):
        self.gallery_view = QListWidget()
        self.gallery_view.setViewMode(QListWidget.ViewMode.IconMode)
        self.gallery_view.setResizeMode(QListWidget.ResizeMode.Adjust)
        self.gallery_view.setWordWrap(True)
        self.gallery_view.setSpacing(20)
        self.gallery_view.itemDoubleClicked.connect(self.show_book_details_from_gallery)
        self.stacked_widget.addWidget(self.gallery_view)

    def toggle_view(self):
        """Chuyển đổi giữa các chế độ xem: Bảng <-> Lưới."""
        current_index = self.stacked_widget.currentIndex()
        # Nếu đang ở Dashboard (index 2), quay về Bảng (index 0)
        if current_index == 2:
            self.stacked_widget.setCurrentIndex(0)
            self.view_toggle_button.setText("Xem dạng Lưới")
            self.dashboard_button.setEnabled(True)
        elif current_index == 0:
            self.stacked_widget.setCurrentIndex(1)
            self.view_toggle_button.setText("Xem dạng Bảng")
        else:
            self.stacked_widget.setCurrentIndex(0)
            self.view_toggle_button.setText("Xem dạng Lưới")

    def toggle_dashboard(self):
        """Chuyển sang chế độ xem Dashboard."""
        self.stacked_widget.setCurrentIndex(2) # Index 2 là DashboardWidget
        self.dashboard_widget.refresh_data() # Cập nhật dữ liệu mới nhất
        self.view_toggle_button.setText("Quay lại Thư viện")
        self.dashboard_button.setEnabled(False) # Vô hiệu hóa nút Dashboard khi đang xem nó

    def load_books_from_db(self, books=None):
        if books is None: 
            self.grand_total_stats = get_library_stats()
            sort_order_str = "ASC" if self.sort_order == Qt.SortOrder.AscendingOrder else "DESC"
            books = get_filtered_books(self.active_filters, self.search_input.text(), self.sort_column, sort_order_str)
        
        # Cập nhật cả hai view
        self.book_table_widget.display_books(books)
        self.update_gallery_view(books)
        
        self.update_status_bar(books)

    def update_gallery_view(self, books):
        """Cập nhật dữ liệu cho Gallery View."""
        self.gallery_view.clear()
        for book_dict_row in books:
            book_dict = dict(book_dict_row)
            book_id = book_dict.get('id')
            cover_path = book_dict.get('cover_path', '')
            title = book_dict.get('ten_sach', 'Không có tên')

            item = QListWidgetItem(self.gallery_view)
            item.setData(Qt.ItemDataRole.UserRole, book_id)
            
            thumbnail_path = get_thumbnail_path(book_id, cover_path)
            item_widget = GalleryItemWidget(thumbnail_path, title)
            item_widget.image_label.doubleClicked.connect(lambda path=cover_path: self.show_full_size_image(path))
            
            item.setSizeHint(item_widget.sizeHint())
            self.gallery_view.addItem(item)
            self.gallery_view.setItemWidget(item, item_widget)

    def apply_filters_debounced(self):
        self.filter_timer.start(400)

    def trigger_filter(self):
        sort_order_str = "ASC" if self.sort_order == Qt.SortOrder.AscendingOrder else "DESC"
        books = get_filtered_books(self.active_filters, self.search_input.text(), self.sort_column, sort_order_str)
        self.load_books_from_db(books)

    def start_processing(self):
        if not self.upload_button.isEnabled(): return

        input_dir = "images_input"
        valid_extensions = ['.png', '.jpg', '.jpeg', '.bmp']
        
        try:
            all_files = os.listdir(input_dir)
            image_files = [os.path.join(input_dir, f) for f in all_files if os.path.splitext(f)[1].lower() in valid_extensions]
        except FileNotFoundError:
            self.show_error(f"Thư mục '{input_dir}' không tồn tại.")
            return

        if not image_files:
            QMessageBox.information(self, "Thông báo", f"Không tìm thấy ảnh nào trong thư mục '{input_dir}'.\n\nVui lòng sao chép ảnh bạn muốn xử lý vào đó.")
            return

        self.image_queue.clear()
        self.image_queue.extend(image_files)
        
        self.total_images_in_queue = len(self.image_queue)
        self.upload_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.process_next_in_queue()

    def process_next_in_queue(self):
        if not self.image_queue:
            self.status_label.setText("Hoàn tất xử lý tất cả các ảnh!")
            self.upload_button.setEnabled(True)
            self.progress_bar.setVisible(False)
            self.grand_total_stats = get_library_stats()
            self.update_status_bar()
            return

        current_path = self.image_queue.pop(0)
        current_count = self.total_images_in_queue - len(self.image_queue)
        self.status_label.setText(f"Đang xử lý ảnh {current_count} trên {self.total_images_in_queue}...")
        self.progress_bar.setValue(0)

        self.thread = QThread(self)
        self.worker = AIServiceWorker(current_path)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.on_processing_finished)
        self.worker.error.connect(self.on_error)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.start()

    def update_progress(self, value, message):
        self.progress_bar.setValue(value)
        self.status_label.setText(f"Ảnh {self.total_images_in_queue - len(self.image_queue) + 1}/{self.total_images_in_queue}: {message}")

    def on_error(self, message):
        self.show_error(message)
        self.process_next_in_queue()

    def on_processing_finished(self, data_from_ai, rotation_angle):
        source_path = self.worker.image_path
        
        # Xoay ảnh gốc trước khi xử lý
        if rotation_angle != 0:
            rotate_image_file(source_path, rotation_angle)

        processed_path = self.process_image_file(data_from_ai)
        
        key_map = get_ai_to_db_key_map()
        book_data = {db_key: data_from_ai.get(ai_key) for ai_key, db_key in key_map.items()}
        book_data['cover_path'] = processed_path

        book_id = add_book(book_data)
        if processed_path:
            create_thumbnail(processed_path, book_id)
            
        embed_and_store_book(book_id)
        self.trigger_filter()
        self.process_next_in_queue()

    def process_image_file(self, data):
        path_to_process = self.worker.image_path
        book_title = data.get("Tên sách", f"unknown_book_{datetime.now().strftime('%Y%m%d%H%M%S')}")
        safe_title = "".join(c for c in book_title if c.isalnum() or c in (' ', '.', '_')).rstrip()
        
        if path_to_process and os.path.exists(path_to_process):
            _, extension = os.path.splitext(path_to_process)
            new_filename = f"{safe_title}{extension}"
            destination_path = os.path.join("images_processed", new_filename)
            try:
                shutil.move(path_to_process, destination_path)
                return destination_path
            except Exception as e:
                self.show_error(f"Không thể di chuyển/đổi tên file ảnh: {e}")
        return None

    def update_status_bar(self, filtered_books=None):
        if filtered_books is None:
            stats = get_library_stats()
            self.grand_total_stats = stats
            self.status_bar.showMessage(f"Tổng sách: {stats.get('total_books', 0)} | Tổng trang: {stats.get('total_pages', 0):,} | Đang đọc: {stats.get('total_pages_reading', 0):,} trang | Tổng giá trị: {stats.get('total_price', 0):,.0f} VNĐ")
            return

        filtered_count = len(filtered_books)
        filtered_pages = 0
        filtered_price = 0
        filtered_reading = 0

        for book in filtered_books:
            try:
                filtered_pages += int(book['so_trang'] or 0)
            except (ValueError, TypeError):
                pass
            try:
                filtered_price += int(book['gia_tiki'] or 0)
            except (ValueError, TypeError):
                pass
            if book['status'] and book['status'].startswith('Đang đọc'):
                match = re.search(r'\((\d+)/', book['status'])
                if match:
                    try: filtered_reading += int(match.group(1))
                    except (ValueError, IndexError): continue
        
        gt = self.grand_total_stats
        msg = (f"Hiển thị: {filtered_count}/{gt.get('total_books', 0)} sách | "
               f"Trang: {filtered_pages:,}/{gt.get('total_pages', 0):,} | "
               f"Đang đọc: {filtered_reading:,}/{gt.get('total_pages_reading', 0):,} | "
               f"Giá trị: {filtered_price:,.0f}/{gt.get('total_price', 0):,.0f} VNĐ")
        self.status_bar.showMessage(msg)

    def show_book_details_from_gallery(self, item):
        book_id = item.data(Qt.ItemDataRole.UserRole)
        if book_id is not None:
            self.display_book_dialog(book_id)

    def display_book_dialog(self, book_id):
        book_data = get_book_by_id(book_id)
        if not book_data:
            self.show_error("Không tìm thấy thông tin sách.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(book_data['ten_sach'])
        dialog.setMinimumSize(800, 700)
        
        dialog_layout = QVBoxLayout(dialog)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        dialog_layout.addWidget(scroll_area)

        scroll_content = QWidget()
        scroll_area.setWidget(scroll_content)
        
        form_layout = QFormLayout(scroll_content)
        form_layout.setRowWrapPolicy(QFormLayout.RowWrapPolicy.WrapAllRows)
        
        for key, value in dict(book_data).items():
            if key == 'id': continue
            display_key = key.replace('_', ' ').title()
            
            if key == 'cover_path':
                label = QLabel("Bìa sách (nháy đúp để phóng to):")
                pixmap_label = ClickableLabel()
                if value and os.path.exists(str(value)):
                    pixmap = QPixmap(str(value))
                    pixmap_label.setPixmap(pixmap.scaledToWidth(200, Qt.TransformationMode.SmoothTransformation))
                    pixmap_label.doubleClicked.connect(lambda path=str(value): self.show_full_size_image(path))
                form_layout.addRow(label, pixmap_label)
            elif key in ['review', 'danh_gia_ca_nhan']:
                field = QTextEdit(str(value if value is not None else ""))
                field.setReadOnly(True)
                field.setMinimumHeight(100)
                form_layout.addRow(QLabel(f"{display_key}:"), field)
            else:
                field = QLineEdit(str(value if value is not None else ""))
                field.setReadOnly(True)
                form_layout.addRow(QLabel(f"{display_key}:"), field)
        
        dialog.exec()

    def show_full_size_image(self, image_path):
        if not image_path or not os.path.exists(image_path):
            self.show_error("Không tìm thấy file ảnh.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(os.path.basename(image_path))
        
        pixmap = QPixmap(image_path)
        label = QLabel()
        label.setPixmap(pixmap)
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(label)
        
        layout = QVBoxLayout(dialog)
        layout.addWidget(scroll_area)
        
        dialog.resize(min(pixmap.width() + 40, 1200), min(pixmap.height() + 40, 800))
        dialog.exec()

    def show_error(self, message):
        QMessageBox.critical(self, "Lỗi", message)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.load_banner_images()

    def closeEvent(self, event):
        self.save_settings()
        super().closeEvent(event)

    def save_settings(self):
        settings = QSettings(SETTINGS_ORGANIZATION, SETTINGS_APP_NAME)
        settings.setValue("geometry", self.saveGeometry())
        settings.setValue("windowState", self.saveState())
        settings.setValue("tableHeaderState", self.book_table_widget.table.horizontalHeader().saveState())

    def load_settings(self):
        settings = QSettings(SETTINGS_ORGANIZATION, SETTINGS_APP_NAME)
        geometry = settings.value("geometry")
        if geometry: self.restoreGeometry(geometry)
        
        state = settings.value("windowState")
        if state: self.restoreState(state)

        header_state = settings.value("tableHeaderState")
        if header_state: self.book_table_widget.table.horizontalHeader().restoreState(header_state)
        
        self.load_banner_images()

    def export_selected_to_excel(self):
        selected_rows = sorted(list(set(index.row() for index in self.book_table_widget.table.selectedIndexes())))
        if not selected_rows:
            QMessageBox.information(self, "Thông báo", "Vui lòng chọn ít nhất một hàng để xuất.")
            return
        
        table = self.book_table_widget.table
        db_cols = self.book_table_widget.db_columns
        book_ids = [table.item(row, db_cols.index('id')).data(Qt.ItemDataRole.UserRole) for row in selected_rows]
        books_data = [dict(get_book_by_id(bid)) for bid in book_ids if get_book_by_id(bid)]
        self.export_to_excel(books_data)

    def export_all_to_excel(self):
        sort_order_str = "ASC" if self.sort_order == Qt.SortOrder.AscendingOrder else "DESC"
        books_data = [dict(book) for book in get_filtered_books(self.active_filters, self.search_input.text(), self.sort_column, sort_order_str)]
        self.export_to_excel(books_data)

    def export_to_excel(self, book_data):
        if not book_data:
            QMessageBox.information(self, "Thông báo", "Không có dữ liệu để xuất.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Lưu file Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        self.status_label.setText("Đang chuẩn bị xuất ra file Excel...")
        self.upload_button.setEnabled(False)
        self.export_button.setEnabled(False)

        headers = [h for h in self.book_table_widget.column_headers if h != 'ID']
        db_cols = [c for c in self.book_table_widget.db_columns if c != 'id']

        self.export_thread = QThread(self)
        self.export_worker = ExcelExportWorker(file_path, headers, db_cols, book_data)
        self.export_worker.moveToThread(self.export_thread)

        self.export_thread.started.connect(self.export_worker.run)
        self.export_worker.finished.connect(self.on_export_finished)
        self.export_worker.error.connect(self.on_export_error)

        self.export_worker.finished.connect(self.export_thread.quit)
        self.export_worker.finished.connect(self.export_worker.deleteLater)
        self.export_thread.finished.connect(self.export_thread.deleteLater)

        self.export_thread.start()

    def on_export_finished(self, file_path):
        self.status_label.setText(f"Đã xuất thành công ra file: {os.path.basename(file_path)}")
        self.upload_button.setEnabled(True)
        self.export_button.setEnabled(True)

        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setText("Xuất file Excel thành công!")
        msg_box.setInformativeText(f"File đã được lưu tại:\n{file_path}")
        
        btn_open_file = msg_box.addButton("Mở File", QMessageBox.ButtonRole.ActionRole)
        btn_open_folder = msg_box.addButton("Mở Thư mục", QMessageBox.ButtonRole.ActionRole)
        btn_close = msg_box.addButton("Đóng", QMessageBox.ButtonRole.RejectRole)
        
        msg_box.exec()

        if msg_box.clickedButton() == btn_open_file:
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(file_path)))
        elif msg_box.clickedButton() == btn_open_folder:
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(os.path.abspath(file_path))))

    def on_export_error(self, error_message):
        self.show_error(error_message)
        self.status_label.setText("Xuất file thất bại.")
        self.upload_button.setEnabled(True)
        self.export_button.setEnabled(True)
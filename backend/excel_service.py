# Tên file: backend/excel_service.py
# CHỨC NĂNG: Cung cấp dịch vụ xuất dữ liệu thư viện sách ra file Excel.
# CHANGELOG:
# - 10:28:00 08/07/2026: [NEW] Khởi tạo module dịch vụ xuất Excel. (Antigravity)

import os
from typing import Any
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_to_excel_file(
    file_path: str,
    headers: list[str],
    db_columns: list[str],
    book_data: list[dict[str, Any]],
) -> None:
    """Xuất danh sách sách ra file Excel có định dạng bảng chuyên nghiệp và chèn ảnh bìa.

    Args:
        file_path: Đường dẫn lưu file Excel đầu ra (.xlsx).
        headers: Danh sách tiêu đề cột hiển thị trên Excel.
        db_columns: Danh sách tên cột tương ứng trong database.
        book_data: Danh sách các dictionary chứa dữ liệu sách.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Active worksheet not found"
    ws.title = "Thư viện sách"

    # 1. Tạo tiêu đề lớn
    ws.merge_cells("A1:Q1")  # Bảng có khoảng 17 cột
    title_cell = ws["A1"]
    title_cell.value = "THƯ VIỆN GIA ĐÌNH LÊ HẢI LƯU"
    title_cell.font = Font(size=20, bold=True, color="0000FF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 2. Ghi tiêu đề cột (Hàng 2)
    ws.append(headers)

    # Định dạng tiêu đề cột
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0070C0", end_color="0070C0", fill_type="solid"
    )
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3. Ghi dữ liệu (Từ hàng 3)
    for book in book_data:
        row_data = []
        for col in db_columns:
            if col == "id":
                continue
            # Nếu là cột cover_path, ta ghi chuỗi rỗng để chỉ hiện ảnh
            if col == "cover_path":
                row_data.append("")
            else:
                row_data.append(book.get(col, ""))
        ws.append(row_data)

    # 4. Chèn ảnh và Định dạng Cell
    # Tìm vị trí cột Bìa sách trong Excel (1-based)
    try:
        excel_cover_col_idx = headers.index("Bìa sách") + 1
        cover_col_letter = get_column_letter(excel_cover_col_idx)
        ws.column_dimensions[cover_col_letter].width = 15
    except ValueError:
        cover_col_letter = "A"

    # Tìm vị trí cột Review để chỉnh độ rộng
    try:
        review_col_idx = headers.index("Review") + 1
        ws.column_dimensions[get_column_letter(review_col_idx)].width = 50
    except ValueError:
        pass

    # Định dạng Wrap Text cho toàn bộ vùng dữ liệu
    last_row = len(book_data) + 2
    last_col_letter = get_column_letter(len(headers))
    data_range = ws[f"A3:{last_col_letter}{last_row}"]

    for row in data_range:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Chèn ảnh bìa vào Excel
    for i, book in enumerate(book_data):
        row_num = i + 3  # Hàng dữ liệu bắt đầu từ 3
        ws.row_dimensions[row_num].height = 95
        cover_path = book.get("cover_path")
        if cover_path and os.path.exists(cover_path):
            try:
                img = OpenpyxlImage(cover_path)
                img.height = 120
                img.width = 90
                cell_address = f"{cover_col_letter}{row_num}"
                ws.add_image(img, cell_address)
            except Exception as img_e:
                print(f"Không thể chèn ảnh {cover_path} vào Excel: {img_e}")

    # 5. Tạo Table (Excel Table)
    tab = Table(displayName="BookTable", ref=f"A2:{last_col_letter}{last_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(file_path)
